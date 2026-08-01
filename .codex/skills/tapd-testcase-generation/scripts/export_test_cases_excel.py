"""Export TAPD testcase JSON to the contracted Excel workbook."""

from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Sequence, TypeAlias

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


JsonObject: TypeAlias = dict[str, object]
ExcelRow: TypeAlias = tuple[str, ...]

SHEET_NAME: str = "测试用例"
HEADERS: tuple[str, ...] = (
    "用例编号",
    "用例名称",
    "用例目录",
    "需求ID",
    "用例类型",
    "用例状态",
    "用例等级",
    "所属端/角色/系统",
    "功能模块",
    "前置条件",
    "测试步骤",
    "预期结果",
    "关联需求点",
    "备注",
)
SCALAR_FIELDS: tuple[str, ...] = (
    "case_id",
    "title",
    "directory",
    "requirement_id",
    "case_type",
    "case_status",
    "priority",
    "system_scope",
    "module",
    "precondition",
)
LIST_FIELDS: tuple[str, ...] = (
    "steps",
    "expected_results",
    "requirement_points",
)
COLUMN_WIDTHS: tuple[float, ...] = (
    12,
    42,
    24,
    14,
    12,
    12,
    12,
    24,
    24,
    42,
    52,
    52,
    42,
    30,
)


def read_json_object(path: Path) -> JsonObject:
    try:
        content: str = path.read_text(encoding="utf-8-sig")
    except OSError as error:
        raise ValueError(f"无法读取文件 {path}: {error}") from error
    if not content.strip():
        raise ValueError(f"文件不能为空: {path}")
    try:
        value: object = json.loads(content)
    except json.JSONDecodeError as error:
        raise ValueError(f"JSON 格式错误 {path}: {error}") from error
    if not isinstance(value, dict):
        raise ValueError(f"JSON 根节点必须是对象: {path}")
    return value


def require_string(value: object, field_name: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{field_name} 必须是非空字符串。")
    return value.strip()


def require_string_list(value: object, field_name: str) -> list[str]:
    if not isinstance(value, list) or not value:
        raise ValueError(f"{field_name} 必须是非空数组。")
    return [
        require_string(item, f"{field_name}[{index}]")
        for index, item in enumerate(value)
    ]


def format_numbered_list(values: list[str]) -> str:
    return "\n".join(
        f"{index}. {value}"
        for index, value in enumerate(values, start=1)
    )


def build_excel_row(raw_case: object, index: int) -> ExcelRow:
    field_prefix: str = f"cases[{index}]"
    if not isinstance(raw_case, dict):
        raise ValueError(f"{field_prefix} 必须是对象。")

    scalar_values: list[str] = [
        require_string(raw_case.get(field_name), f"{field_prefix}.{field_name}")
        for field_name in SCALAR_FIELDS
    ]
    list_values: list[str] = [
        format_numbered_list(
            require_string_list(
                raw_case.get(field_name),
                f"{field_prefix}.{field_name}",
            )
        )
        for field_name in LIST_FIELDS
    ]
    remarks: str = require_string(raw_case.get("remarks"), f"{field_prefix}.remarks")
    return tuple([*scalar_values, *list_values, remarks])


def build_excel_rows(payload: JsonObject) -> list[ExcelRow]:
    raw_cases: object = payload.get("cases")
    if not isinstance(raw_cases, list) or not raw_cases:
        raise ValueError("cases 必须是非空数组。")
    raw_total_count: object = payload.get("total_count")
    if isinstance(raw_total_count, bool) or not isinstance(raw_total_count, int):
        raise ValueError("total_count 必须是整数。")
    if raw_total_count != len(raw_cases):
        raise ValueError("total_count 与 cases 数量不一致。")
    return [
        build_excel_row(raw_case, index)
        for index, raw_case in enumerate(raw_cases)
    ]


def style_worksheet(worksheet: Worksheet, row_count: int) -> None:
    header_fill: PatternFill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font: Font = Font(color="FFFFFF", bold=True)
    header_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    body_alignment: Alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
        for cell in row:
            cell.alignment = body_alignment

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:N{row_count + 1}"
    worksheet.row_dimensions[1].height = 24


def save_workbook(rows: list[ExcelRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook: Workbook = Workbook()
    worksheet: Worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    style_worksheet(worksheet, len(rows))

    descriptor: int
    temporary_name: str
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=".xlsx",
        dir=output_path.parent,
    )
    os.close(descriptor)
    temporary_path: Path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    except OSError as error:
        raise ValueError(f"无法写入 Excel 文件 {output_path}: {error}") from error
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def export_test_cases(output_directory: Path) -> Path:
    input_path: Path = output_directory / "tapd_cases.json"
    output_path: Path = output_directory / "test_cases.xlsx"
    payload: JsonObject = read_json_object(input_path)
    rows: list[ExcelRow] = build_excel_rows(payload)
    save_workbook(rows, output_path)
    return output_path


def parse_arguments(argv: Sequence[str]) -> argparse.Namespace:
    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        description="Export tapd_cases.json to test_cases.xlsx."
    )
    parser.add_argument("output_directory", type=Path)
    return parser.parse_args(argv)


def main(argv: Sequence[str]) -> int:
    arguments: argparse.Namespace = parse_arguments(argv)
    try:
        output_path: Path = export_test_cases(arguments.output_directory)
    except ValueError as error:
        print(f"Excel export failed: {error}", file=sys.stderr)
        return 1
    print(f"Excel export passed: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
